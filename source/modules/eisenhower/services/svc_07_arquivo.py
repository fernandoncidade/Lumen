from PySide6.QtCore import QCoreApplication
from PySide6.QtWidgets import QFileDialog, QMessageBox, QListWidgetItem
from PySide6.QtCore import Qt
import os
from source.utils.LogManager import LogManager
logger = LogManager.get_logger()

def get_text(text):
    return QCoreApplication.translate("App", text)

def _clear_all_lists(app):
    for lst in (
        app.quadrant1_list, app.quadrant2_list, app.quadrant3_list, app.quadrant4_list,
        app.quadrant1_completed_list, app.quadrant2_completed_list, app.quadrant3_completed_list, app.quadrant4_completed_list
    ):
        lst.clear()

def _add_placeholders_after_clear(app):
    app.add_placeholder(app.quadrant1_list, get_text("1º Quadrante"))
    app.add_placeholder(app.quadrant2_list, get_text("2º Quadrante"))
    app.add_placeholder(app.quadrant3_list, get_text("3º Quadrante"))
    app.add_placeholder(app.quadrant4_list, get_text("4º Quadrante"))
    app.add_placeholder(app.quadrant1_completed_list, get_text("Nenhuma Tarefa Concluída"))
    app.add_placeholder(app.quadrant2_completed_list, get_text("Nenhuma Tarefa Concluída"))
    app.add_placeholder(app.quadrant3_completed_list, get_text("Nenhuma Tarefa Concluída"))
    app.add_placeholder(app.quadrant4_completed_list, get_text("Nenhuma Tarefa Concluída"))

def novo(app):
    _clear_all_lists(app)
    _add_placeholders_after_clear(app)
    app.task_input.clear()
    try:
        app.save_tasks()
        if hasattr(app, "calendar_pane") and app.calendar_pane:
            app.calendar_pane.calendar_panel.update_task_list()

    except Exception as e:
        logger.error(f"Erro ao iniciar nova sessão: {e}", exc_info=True)

    QMessageBox.information(app, get_text("Novo"), get_text("Nova sessão iniciada.") or "Nova sessão iniciada.")

def limpar_tudo(app):
    _clear_all_lists(app)
    _add_placeholders_after_clear(app)
    app.task_input.clear()
    try:
        app.save_tasks()
        if hasattr(app, "calendar_pane") and app.calendar_pane:
            app.calendar_pane.calendar_panel.update_task_list()

    except Exception as e:
        logger.error(f"Erro ao limpar tudo: {e}", exc_info=True)

    QMessageBox.information(app, get_text("Limpar"), get_text("Todos os dados foram removidos.") or "Todos os dados foram removidos.")

def sair(app):
    from PySide6.QtWidgets import QApplication
    QApplication.quit()

def abrir_arquivo(app):
    path, filt = QFileDialog.getOpenFileName(app, get_text("Abrir"), os.path.expanduser("~"), "Excel (*.xlsx);;PDF (*.pdf)")
    if not path:
        return

    def _parse_text_date_time(raw_text: str):
        base = (raw_text or "").strip()
        date_iso = None
        time_str = None
        if " — " in base:
            left, right = base.rsplit(" — ", 1)
            tail = right.strip()
            try:
                from PySide6.QtCore import QDate, QTime
                fmts = []
                try:
                    fmts.append(app.date_input.displayFormat())

                except Exception as e:
                    logger.error(f"Erro ao obter formato de data para parsing: {e}", exc_info=True)

                fmts += ["dd/MM/yyyy", "d/M/yyyy", "MM/dd/yyyy", "M/d/yyyy", "yyyy-MM-dd"]

                parts = tail.split()
                if len(parts) == 2:
                    d_candidate, t_candidate = parts
                    for fmt in fmts:
                        qd = QDate.fromString(d_candidate, fmt)
                        if qd and qd.isValid():
                            qt = QTime.fromString(t_candidate, "HH:mm")
                            if qt.isValid():
                                date_iso = qd.toString(Qt.ISODate)
                                time_str = qt.toString("HH:mm")
                                base = left.strip()
                                break

                if date_iso is None:
                    for fmt in fmts:
                        qd = QDate.fromString(tail, fmt)
                        if qd and qd.isValid():
                            date_iso = qd.toString(Qt.ISODate)
                            base = left.strip()
                            break

                if date_iso is None:
                    from PySide6.QtCore import QTime
                    qt = QTime.fromString(tail, "HH:mm")
                    if qt.isValid():
                        time_str = qt.toString("HH:mm")
                        base = left.strip()

            except Exception as e:
                logger.error(f"Erro ao fazer parsing de data/hora: {e}", exc_info=True)

        return base, date_iso, time_str

    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook

        except Exception as e:
            logger.error(f"openpyxl não está disponível: {e}", exc_info=True)
            QMessageBox.critical(app, get_text("Erro"), get_text("openpyxl não está disponível."))
            return

        try:
            wb = load_workbook(path, read_only=True)
            def populate_from_sheet(name, lst, completed=False):
                if name in wb.sheetnames:
                    sheet = wb[name]
                    lst.clear()
                    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
                        val = row[0]
                        if val and str(val).strip():
                            raw_text = str(val).strip()
                            text, date_iso, time_str = _parse_text_date_time(raw_text)
                            item = QListWidgetItem(raw_text)
                            item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                            item.setCheckState(Qt.Checked if completed else Qt.Unchecked)
                            item.setData(Qt.UserRole, {"text": text, "date": date_iso, "time": time_str})
                            tooltip_lines = []
                            if date_iso:
                                try:
                                    from PySide6.QtCore import QDate
                                    qd = QDate.fromString(date_iso, Qt.ISODate)
                                    if qd.isValid():
                                        tooltip_lines.append(f"{get_text('Data') or 'Data'}: {qd.toString(app.date_input.displayFormat())}")

                                except Exception as e:
                                    logger.error(f"Erro ao definir tooltip de data: {e}", exc_info=True)

                            if time_str:
                                tooltip_lines.append(f"{get_text('Horário') or 'Horário'}: {time_str}")

                            if tooltip_lines:
                                item.setToolTip("\n".join(tooltip_lines))

                            try:
                                app.insert_task_into_quadrant_list(lst, item)

                            except Exception as e:
                                logger.error(f"Erro ao inserir tarefa na lista do quadrante: {e}", exc_info=True)
                                lst.addItem(item)

            populate_from_sheet("quadrant1", app.quadrant1_list, completed=False)
            populate_from_sheet("quadrant1_completed", app.quadrant1_completed_list, completed=True)
            populate_from_sheet("quadrant2", app.quadrant2_list, completed=False)
            populate_from_sheet("quadrant2_completed", app.quadrant2_completed_list, completed=True)
            populate_from_sheet("quadrant3", app.quadrant3_list, completed=False)
            populate_from_sheet("quadrant3_completed", app.quadrant3_completed_list, completed=True)
            populate_from_sheet("quadrant4", app.quadrant4_list, completed=False)
            populate_from_sheet("quadrant4_completed", app.quadrant4_completed_list, completed=True)

            app.save_tasks()
            try:
                if hasattr(app, "calendar_pane") and app.calendar_pane:
                    app.calendar_pane.calendar_panel.update_task_list()

            except Exception as e:
                logger.error(f"Erro ao atualizar lista de tarefas no calendário: {e}", exc_info=True)

            QMessageBox.information(app, get_text("Abrir"), get_text("Arquivo importado com sucesso."))

        except Exception as e:
            logger.error(f"Erro ao importar arquivo XLSX: {e}", exc_info=True)
            QMessageBox.critical(app, get_text("Erro"), f"{get_text('Erro') or 'Erro'}: {e}")

    elif ext == ".pdf":
        try:
            from PyPDF2 import PdfReader

        except Exception as e:
            logger.error(f"PyPDF2 não está disponível: {e}", exc_info=True)
            QMessageBox.critical(app, get_text("Erro"), get_text("PyPDF2 não está disponível para ler PDF."))
            return

        try:
            reader = PdfReader(path)
            text = ""
            for page in reader.pages:
                try:
                    page_text = page.extract_text() or ""
                    text += page_text + "\n"

                except Exception as e:
                    logger.error(f"Erro ao extrair texto da página do PDF: {e}", exc_info=True)

            import unicodedata, re
            def normalize(s: str) -> str:
                s2 = unicodedata.normalize("NFKD", s)
                s2 = s2.encode("ASCII", "ignore").decode("ASCII")
                s2 = s2.lower()
                s2 = re.sub(r'\s+', ' ', s2).strip()
                return s2

            title_map = {
                normalize("1º Quadrante - Importante e Urgente"): "quadrant1",
                normalize("Concluídas 1º Quadrante"): "quadrant1_completed",
                normalize("2º Quadrante - Importante, mas Não Urgente"): "quadrant2",
                normalize("Concluídas 2º Quadrante"): "quadrant2_completed",
                normalize("3º Quadrante - Não Importante, mas Urgente"): "quadrant3",
                normalize("Concluídas 3º Quadrante"): "quadrant3_completed",
                normalize("4º Quadrante - Não Importante e Não Urgente"): "quadrant4",
                normalize("Concluídas 4º Quadrante"): "quadrant4_completed",
            }

            segments = {k: [] for k in title_map.values()}
            current_key = None
            for raw_line in text.splitlines():
                line = raw_line.strip()
                if not line:
                    continue

                n = normalize(line)
                if n in title_map:
                    current_key = title_map[n]
                    continue

                if current_key:
                    cleaned = re.sub(r'^[\-\u2013\u2014\u2022\u2023\•\*\•\s]+', '', line).strip()
                    if cleaned:
                        segments[current_key].append(cleaned)

            if not any(segments.values()):
                lower = text.lower()
                keys = ["quadrant1", "quadrant1_completed", "quadrant2", "quadrant2_completed",
                        "quadrant3", "quadrant3_completed", "quadrant4", "quadrant4_completed"]

                found_any = False
                for i, key in enumerate(keys):
                    start = lower.find(key + ":")
                    if start != -1:
                        found_any = True
                        end = len(lower)
                        for k in keys[i+1:]:
                            j = lower.find(k + ":", start+1)
                            if j != -1:
                                end = j
                                break

                        segment_text = text[start+len(key)+1:end].strip()
                        for line in segment_text.splitlines():
                            line = line.strip()
                            if line:
                                cleaned = re.sub(r'^[\-\u2013\u2014\u2022\u2023\•\*\s]+', '', line).strip()
                                if cleaned:
                                    segments[key].append(cleaned)

                if not found_any:
                    QMessageBox.warning(app, get_text("Abrir"), get_text("PDF não está no formato compatível."))
                    return

            def populate_from_list(key, lst, completed=False):
                lst.clear()
                items = segments.get(key, [])
                for it in items:
                    raw_text = it.strip()
                    text_only, date_iso, time_str = _parse_text_date_time(raw_text)
                    item = QListWidgetItem(raw_text)  # mantém exibição
                    item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    item.setCheckState(Qt.Checked if completed else Qt.Unchecked)
                    item.setData(Qt.UserRole, {"text": text_only, "date": date_iso, "time": time_str})
                    tooltip_lines = []
                    if date_iso:
                        try:
                            from PySide6.QtCore import QDate
                            qd = QDate.fromString(date_iso, Qt.ISODate)
                            if qd.isValid():
                                tooltip_lines.append(f"{get_text('Data') or 'Data'}: {qd.toString(app.date_input.displayFormat())}")

                        except Exception as e:
                            logger.error(f"Erro ao definir tooltip de data no PDF: {e}", exc_info=True)

                    if time_str:
                        tooltip_lines.append(f"{get_text('Horário') or 'Horário'}: {time_str}")

                    if tooltip_lines:
                        item.setToolTip("\n".join(tooltip_lines))

                    try:
                        app.insert_task_into_quadrant_list(lst, item)

                    except Exception as e:
                        logger.error(f"Erro ao inserir tarefa na lista do quadrante (PDF): {e}", exc_info=True)
                        lst.addItem(item)

            populate_from_list("quadrant1", app.quadrant1_list, False)
            populate_from_list("quadrant1_completed", app.quadrant1_completed_list, True)
            populate_from_list("quadrant2", app.quadrant2_list, False)
            populate_from_list("quadrant2_completed", app.quadrant2_completed_list, True)
            populate_from_list("quadrant3", app.quadrant3_list, False)
            populate_from_list("quadrant3_completed", app.quadrant3_completed_list, True)
            populate_from_list("quadrant4", app.quadrant4_list, False)
            populate_from_list("quadrant4_completed", app.quadrant4_completed_list, True)

            app.save_tasks()
            try:
                if hasattr(app, "calendar_pane") and app.calendar_pane:
                    app.calendar_pane.calendar_panel.update_task_list()

            except Exception as e:
                logger.error(f"Erro ao atualizar lista de tarefas no calendário (PDF): {e}", exc_info=True)

            QMessageBox.information(app, get_text("Abrir"), get_text("PDF importado com sucesso."))

        except Exception as e:
            logger.error(f"Erro ao importar arquivo PDF: {e}", exc_info=True)
            QMessageBox.critical(app, get_text("Erro"), f"{get_text('Erro') or 'Erro'}: {e}")

    else:
        QMessageBox.warning(app, get_text("Abrir"), get_text("Formato de arquivo não suportado."))

def salvar_como(app):
    path, filt = QFileDialog.getSaveFileName(app, get_text("Salvar"), os.path.expanduser("~"), "Excel (*.xlsx);;PDF (*.pdf)")
    if not path:
        return

    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        try:
            from openpyxl import Workbook

        except Exception as e:
            logger.error(f"openpyxl não está disponível para salvar XLSX: {e}", exc_info=True)
            QMessageBox.critical(app, get_text("Erro"), get_text("openpyxl não está disponível para salvar XLSX."))
            return

        try:
            wb = Workbook()
            def write_sheet(name, lst):
                if name in wb.sheetnames:
                    ws = wb[name]
                else:
                    ws = wb.create_sheet(title=name)

                for i, r in enumerate(list(ws.rows)):
                    pass

                row = 1
                for i in range(lst.count()):
                    item = lst.item(i)
                    if item and (item.flags() & Qt.ItemIsSelectable):
                        ws.cell(row=row, column=1, value=item.text())
                        row += 1

            if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
                ws_default = wb["Sheet"]
                wb.remove(ws_default)

            write_sheet("quadrant1", app.quadrant1_list)
            write_sheet("quadrant1_completed", app.quadrant1_completed_list)
            write_sheet("quadrant2", app.quadrant2_list)
            write_sheet("quadrant2_completed", app.quadrant2_completed_list)
            write_sheet("quadrant3", app.quadrant3_list)
            write_sheet("quadrant3_completed", app.quadrant3_completed_list)
            write_sheet("quadrant4", app.quadrant4_list)
            write_sheet("quadrant4_completed", app.quadrant4_completed_list)

            wb.save(path)
            QMessageBox.information(app, get_text("Salvar"), get_text("Arquivo salvo com sucesso."))

        except Exception as e:
            logger.error(f"Erro ao salvar arquivo XLSX: {e}", exc_info=True)
            QMessageBox.critical(app, get_text("Erro"), f"{get_text('Erro') or 'Erro'}: {e}")

    elif ext == ".pdf":
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas

        except Exception as e:
            logger.error(f"reportlab não está disponível para salvar PDF: {e}", exc_info=True)
            QMessageBox.critical(app, get_text("Erro"), get_text("reportlab não está disponível para salvar PDF."))
            return

        try:
            c = canvas.Canvas(path, pagesize=A4)
            width, height = A4
            x_margin = 40
            y = height - 40
            c.setFont("Helvetica-Bold", 14)
            c.drawString(x_margin, y, "EISENHOWER ORGANIZER")
            y -= 30
            c.setFont("Helvetica-Bold", 12)
            sections = [
                ("1º Quadrante - Importante e Urgente", app.quadrant1_list),
                ("Concluídas 1º Quadrante", app.quadrant1_completed_list),
                ("2º Quadrante - Importante, mas Não Urgente", app.quadrant2_list),
                ("Concluídas 2º Quadrante", app.quadrant2_completed_list),
                ("3º Quadrante - Não Importante, mas Urgente", app.quadrant3_list),
                ("Concluídas 3º Quadrante", app.quadrant3_completed_list),
                ("4º Quadrante - Não Importante e Não Urgente", app.quadrant4_list),
                ("Concluídas 4º Quadrante", app.quadrant4_completed_list),
            ]
            c.setFont("Helvetica", 10)
            for title, lst in sections:
                if y < 80:
                    c.showPage()
                    y = height - 40
                    c.setFont("Helvetica", 10)

                c.setFont("Helvetica-Bold", 11)
                c.drawString(x_margin, y, title)
                y -= 16
                c.setFont("Helvetica", 10)
                if lst.count() == 0:
                    c.drawString(x_margin + 10, y, "-")
                    y -= 12

                else:
                    for i in range(lst.count()):
                        item = lst.item(i)
                        if item and (item.flags() & Qt.ItemIsSelectable):
                            c.drawString(x_margin + 10, y, f"- {item.text()}")
                            y -= 12
                            if y < 60:
                                c.showPage()
                                y = height - 40
                                c.setFont("Helvetica", 10)

                y -= 8

            c.save()
            QMessageBox.information(app, get_text("Salvar"), get_text("PDF salvo com sucesso."))

        except Exception as e:
            logger.error(f"Erro ao salvar arquivo PDF: {e}", exc_info=True)
            QMessageBox.critical(app, get_text("Erro"), f"{get_text('Erro') or 'Erro'}: {e}")

    else:
        QMessageBox.warning(app, get_text("Salvar"), get_text("Extensão não suportada."))
